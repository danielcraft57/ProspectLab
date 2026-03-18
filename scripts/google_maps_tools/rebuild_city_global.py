import argparse
from pathlib import Path
from typing import List


def _safe_filename(name: str) -> str:
    keep: List[str] = []
    for ch in name.strip():
        if ch.isalnum():
            keep.append(ch)
        elif ch in {" ", "-", "_"}:
            keep.append("_")
    out = "".join(keep)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "ville"


EXPORT_COLUMNS = [
    "name",
    "website",
    "category",
    "phone_number",
    "country",
    "address_1",
    "address_2",
    "longitude",
    "latitude",
    "rating",
    "reviews_count",
    "category_translate",
]


def _load_rows_from_xlsx(path: Path) -> List[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl n'est pas installe. Installe-le avec: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []

        header_map = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}
        out: List[dict] = []
        for row in rows_iter:
            obj = {}
            for col in EXPORT_COLUMNS:
                idx = header_map.get(col)
                obj[col] = row[idx] if idx is not None and idx < len(row) else None
            out.append(obj)
        return out
    finally:
        wb.close()


def _write_xlsx(path: Path, rows: List[dict]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl n'est pas installe. Installe-le avec: pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "export"
    ws.append(EXPORT_COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in EXPORT_COLUMNS])
    wb.save(path)


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Reconstruit le fichier global Data/<Ville>/<Ville>.xlsx a partir des xlsx existants (groupes/secteurs)."
    )
    p.add_argument("--city", required=True, help="Ville (ex: Strasbourg)")
    p.add_argument("--out-dir", required=True, help="Dossier racine (ex: Data)")
    p.add_argument(
        "--sources",
        choices=["auto", "groups-only", "all-xlsx"],
        default="auto",
        help="auto: prefere fichiers de groupes; groups-only: uniquement xlsx racine; all-xlsx: tous les xlsx de la ville",
    )
    args = p.parse_args(argv)

    city = args.city.strip()
    out_root = Path(args.out_dir).expanduser().resolve()
    out_city_dir = out_root / city
    if not out_city_dir.exists():
        raise SystemExit(f"Dossier ville introuvable: {out_city_dir}")

    out_all = out_city_dir / f"{_safe_filename(city)}.xlsx"

    xlsx_all = sorted([p for p in out_city_dir.rglob("*.xlsx") if p.is_file()])
    xlsx_all = [p for p in xlsx_all if p.name != out_all.name]

    root_xlsx = sorted([p for p in out_city_dir.glob("*.xlsx") if p.is_file()])
    root_xlsx = [p for p in root_xlsx if p.name != out_all.name]

    if args.sources == "groups-only":
        sources = root_xlsx
    elif args.sources == "all-xlsx":
        sources = xlsx_all
    else:
        # auto: si on a des fichiers xlsx a la racine, on suppose que ce sont les groupes et on les prefere
        sources = root_xlsx or xlsx_all

    merged_rows: List[dict] = []
    for fpath in sources:
        merged_rows.extend(_load_rows_from_xlsx(fpath))

    seen = set()
    deduped: List[dict] = []
    for r in merged_rows:
        key = (
            str(r.get("website") or "").strip().lower(),
            str(r.get("name") or "").strip().lower(),
            str(r.get("address_1") or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)

    _write_xlsx(out_all, deduped)
    print(f"Fichier global reconstruit: {out_all} (brut: {len(merged_rows)}, apres dedup: {len(deduped)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

