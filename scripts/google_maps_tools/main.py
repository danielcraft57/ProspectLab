import argparse
import json
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests


@dataclass(frozen=True)
class Preset:
    city: str
    region: str
    department: str
    lat: float
    lng: float
    levels: Dict[str, Dict[str, Any]]


PLACES_TEXTSEARCH_URL = "https://maps.googleapis.com/maps/api/place/textsearch/json"


def _safe_filename(name: str) -> str:
    keep: List[str] = []
    for ch in name.strip():
        if ch.isalnum():
            keep.append(ch)
        elif ch in {" ", "-", "_"}:
            keep.append("_")
    out = "".join(keep)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "ville"


def _load_presets(path: Path) -> List[Preset]:
    data = json.loads(path.read_text(encoding="utf-8"))
    presets_raw = data.get("presets") or []
    out: List[Preset] = []
    for p in presets_raw:
        if not isinstance(p, dict):
            continue
        city = str(p.get("city") or "").strip()
        region = str(p.get("region") or "").strip()
        department = str(p.get("department") or "").strip()
        levels = p.get("levels") or {}
        if not city or not region:
            continue
        try:
            lat = float(p.get("lat"))
            lng = float(p.get("lng"))
        except Exception:
            continue
        if not isinstance(levels, dict) or not levels:
            continue
        out.append(Preset(city=city, region=region, department=department, lat=lat, lng=lng, levels=levels))
    return out


def _city_dir(out_root: Path, preset: Preset) -> Path:
    region = (preset.region or "-").strip() or "-"
    department = (preset.department or "-").strip() or "-"
    return out_root / region / department / preset.city


def _status_for_preset(out_root: Path, preset: Preset) -> Tuple[str, str]:
    now_ts = __import__("time").time()
    recent_window_s = 20 * 60
    city_dir = _city_dir(out_root, preset)
    xlsx = city_dir / f"{_safe_filename(preset.city)}.xlsx"
    if not city_dir.exists():
        return "PAS FAIT", "red"

    cache_file = city_dir / "places_details_cache.jsonl"
    sector_xlsx = [p for p in city_dir.glob("*.xlsx") if p.name.lower() != xlsx.name.lower()]

    def _is_recent(path: Path) -> bool:
        try:
            return (now_ts - path.stat().st_mtime) <= recent_window_s
        except Exception:
            return False

    recent_activity = False
    if cache_file.exists() and _is_recent(cache_file):
        recent_activity = True
    if xlsx.exists() and _is_recent(xlsx):
        recent_activity = True
    if not recent_activity:
        recent_activity = any(_is_recent(p) for p in sector_xlsx[:10])

    if not xlsx.exists():
        if sector_xlsx or cache_file.exists():
            return ("EN COURS", "cyan") if recent_activity else ("INCOMPLET", "yellow")
        return "PAS FAIT", "red"
    try:
        sz = xlsx.stat().st_size
    except Exception:
        return ("EN COURS", "cyan") if recent_activity else ("FAIT", "green")
    if sz < 5_000:
        return ("EN COURS", "cyan") if recent_activity else ("INCOMPLET", "yellow")
    if recent_activity:
        return "EN COURS", "cyan"
    return "FAIT", "green"


def _api_key_file(here: Path) -> Path:
    return here / ".google_maps_api_key"


def _read_file_api_key(here: Path) -> str:
    path = _api_key_file(here)
    if not path.exists():
        return ""
    try:
        return path.read_text(encoding="utf-8").strip()
    except Exception:
        return ""


def _write_file_api_key(here: Path, api_key: str) -> None:
    path = _api_key_file(here)
    path.write_text(api_key.strip() + "\n", encoding="utf-8")


def _resolve_api_key(here: Path) -> Tuple[str, str]:
    file_key = _read_file_api_key(here)
    if file_key:
        return file_key, str(_api_key_file(here))
    return "", "aucune"


def _validate_api_key(api_key: str, timeout_s: int = 8) -> Tuple[bool, str]:
    params = {"key": api_key, "query": "boulangerie Metz", "language": "fr", "region": "fr"}
    try:
        response = requests.get(PLACES_TEXTSEARCH_URL, params=params, timeout=timeout_s)
        response.raise_for_status()
        data = response.json()
    except Exception as exc:
        return False, f"Erreur reseau: {exc}"

    status = str(data.get("status") or "")
    if status in {"OK", "ZERO_RESULTS"}:
        return True, "Cle API valide."
    message = str(data.get("error_message") or "Aucun detail fourni par Google.")
    return False, f"Cle invalide ({status}): {message}"


def _ensure_valid_api_key(console, here: Path) -> str:
    from rich.panel import Panel

    while True:
        api_key, source = _resolve_api_key(here)
        if not api_key:
            console.print(
                Panel.fit(
                    "Aucune cle Google Maps detectee.\n"
                    "Colle une cle API Places pour continuer.",
                    style="bold yellow",
                )
            )
            new_key = console.input("Nouvelle cle API: ").strip()
            if not new_key:
                raise SystemExit("Cle vide. Arret.")
            _write_file_api_key(here, new_key)
            continue

        console.print(f"[dim]Cle API chargee depuis: {source}[/]")
        ok, msg = _validate_api_key(api_key=api_key)
        if ok:
            console.print(f"[green]{msg}[/]")
            return api_key

        console.print(Panel.fit(msg, style="bold red"))
        choice = console.input("Modifier la cle maintenant ? (o/n) [o] ").strip().lower() or "o"
        if choice not in {"o", "oui", "y", "yes"}:
            raise SystemExit("Cle API invalide. Arret.")
        new_key = console.input("Nouvelle cle API: ").strip()
        if not new_key:
            console.print("Cle vide, reessaie.", style="yellow")
            continue
        _write_file_api_key(here, new_key)


def _require_rich():
    try:
        from rich.console import Console  # noqa: F401
    except Exception as e:
        msg = (
            "Ce script utilise une interface TUI (tableaux/couleurs) via 'rich'.\n"
            "Installe-le avec:\n\n"
            "  pip install rich\n"
        )
        raise SystemExit(msg) from e


def _build_command(
    *,
    here: Path,
    out_root: Path,
    groups_file: Optional[Path],
    preset: Preset,
    level: str,
    mode: str,
    language: str,
    region_code: Optional[str],
    timeout_s: int,
    limit_per_sector: int,
    details_max: int,
    cache_path: Optional[Path],
    category_translate: str,
) -> List[str]:
    script = here / "export_city_sectors.py"
    if not script.exists():
        raise SystemExit(f"Script introuvable: {script}")

    lvl = preset.levels.get(level)
    if level == "tres_large":
        base = preset.levels.get("large") or preset.levels.get("agglo") or preset.levels.get("centre")
        if not isinstance(base, dict):
            raise SystemExit("Aucun level de base disponible pour calculer 'tres_large'.")
        lvl = {
            "radius": int(base.get("radius", 4000)) * 2,
            "grid_step_m": int(base.get("grid_step_m", 4000)) * 2,
            "grid_rings": max(4, int(base.get("grid_rings", 3)) + 1),
        }
    if not isinstance(lvl, dict):
        raise SystemExit(f"Level inconnu: {level}")

    cmd = [
        sys.executable,
        str(script),
        "--city",
        preset.city,
        "--out-dir",
        str(out_root),
        "--mode",
        mode,
        "--language",
        language,
        "--timeout",
        str(int(timeout_s)),
        "--limit-per-sector",
        str(int(limit_per_sector)),
        "--details-max",
        str(int(details_max)),
        "--category-translate",
        category_translate,
    ]

    if groups_file:
        cmd.extend(["--groups-file", str(groups_file)])

    if region_code:
        cmd.extend(["--region", str(region_code)])

    if cache_path:
        cmd.extend(["--cache", str(cache_path)])

    if mode == "nearby-grid":
        cmd.extend(
            [
                "--lat",
                str(float(preset.lat)),
                "--lng",
                str(float(preset.lng)),
                "--radius",
                str(int(lvl.get("radius"))),
                "--grid-step-m",
                str(int(lvl.get("grid_step_m"))),
                "--grid-rings",
                str(int(lvl.get("grid_rings"))),
            ]
        )
    return cmd


def _pick_index_numbered(console, title: str, items: List[str], *, allow_back: bool) -> Optional[int]:
    from rich.panel import Panel

    while True:
        console.print(Panel.fit(title, style="bold cyan"))
        for i, label in enumerate(items, start=1):
            console.print(f"[bold]{i:>2}[/]. {label}")
        console.print()
        hint = "Choix: [bold]numero[/]"
        if allow_back:
            hint += " | [bold]b[/]=retour"
        hint += " | [bold]q[/]=quitter"
        console.print(hint, style="dim")
        raw = console.input("> ").strip().lower()
        if raw == "q":
            raise SystemExit(0)
        if allow_back and raw == "b":
            return None
        try:
            n = int(raw)
        except Exception:
            console.print("Entrée invalide. Recommence.", style="yellow")
            continue
        if 1 <= n <= len(items):
            return n - 1
        console.print("Hors plage. Recommence.", style="yellow")


def main(argv: Optional[List[str]] = None) -> int:
    _require_rich()
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table

    here = Path(__file__).resolve().parent
    default_presets = here / "presets_grand_est.json"
    default_groups = here / "groupes_secteurs_grands.json"

    p = argparse.ArgumentParser(description="Interface TUI: region -> ville -> level -> export.")
    p.add_argument("--presets", default=str(default_presets), help="Fichier presets JSON")
    p.add_argument("--out-dir", default="Data", help="Dossier racine de sortie")
    p.add_argument("--groups-file", default=str(default_groups), help="JSON de groupes")
    p.add_argument("--no-groups", action="store_true", help="Ne pas utiliser les groupes")
    p.add_argument("--mode", default="nearby-grid", choices=["nearby-grid", "text"], help="Mode de recherche")
    p.add_argument(
        "--level",
        default="agglo",
        choices=["centre", "agglo", "large", "tres_large"],
        help="Level par défaut",
    )
    p.add_argument("--language", default="fr", help="Langue Places")
    p.add_argument("--region-code", default="fr", help="Code region Places")
    p.add_argument("--timeout", type=int, default=25, help="Timeout reseau (s)")
    p.add_argument("--limit-per-sector", type=int, default=200, help="Limit export / secteur")
    p.add_argument("--details-max", type=int, default=200, help="Budget details max / secteur")
    p.add_argument("--cache", help="Cache JSONL (defaut: Data/<Ville>/places_details_cache.jsonl)")
    p.add_argument(
        "--category-translate",
        default="builtin",
        choices=["none", "builtin"],
        help="Traduction de category (defaut: builtin)",
    )
    p.add_argument("--search", help="Filtre texte (region/ville/departement)")
    p.add_argument("--dry-run", action="store_true", help="Affiche la commande sans executer")
    args = p.parse_args(argv)

    console = Console()
    console.print(Panel.fit("ProspectLab - Export Google Maps", style="bold white on blue"))
    _ensure_valid_api_key(console, here)

    presets_path = Path(str(args.presets)).expanduser().resolve()
    if not presets_path.exists():
        raise SystemExit(f"Presets introuvable: {presets_path}")

    out_root = Path(str(args.out_dir)).expanduser().resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    groups_file = None if args.no_groups else Path(str(args.groups_file)).expanduser().resolve()
    if groups_file and not groups_file.exists():
        raise SystemExit(f"Groups file introuvable: {groups_file}")

    presets = _load_presets(presets_path)
    if args.search:
        q = args.search.strip().lower()
        presets = [
            x
            for x in presets
            if q in x.city.lower() or q in x.region.lower() or (x.department and q in x.department.lower())
        ]
    if not presets:
        raise SystemExit("Aucun preset disponible (après filtre).")

    presets_by_region: Dict[str, List[Preset]] = {}
    for pr in presets:
        presets_by_region.setdefault(pr.region, []).append(pr)
    for r in presets_by_region:
        presets_by_region[r] = sorted(presets_by_region[r], key=lambda x: x.city.lower())

    regions = sorted(presets_by_region.keys(), key=str.lower)
    region_labels: List[str] = []
    for r in regions:
        deps = sorted({(x.department or "-").strip() for x in presets_by_region[r]}, key=str.lower)
        region_labels.append(f"[bold]{r}[/]  [dim]({len(presets_by_region[r])} villes, {len(deps)} departements)[/]")
    ridx = _pick_index_numbered(console, "1) Choisis une région", region_labels, allow_back=False)
    if ridx is None:
        return 0
    region = regions[ridx]

    region_presets = presets_by_region[region]
    presets_by_department: Dict[str, List[Preset]] = {}
    for pr in region_presets:
        dep = (pr.department or "-").strip()
        presets_by_department.setdefault(dep, []).append(pr)
    for d in presets_by_department:
        presets_by_department[d] = sorted(presets_by_department[d], key=lambda x: x.city.lower())

    departments = sorted(presets_by_department.keys(), key=str.lower)
    department_labels = [f"[bold]{d}[/]  [dim]({len(presets_by_department[d])} villes)[/]" for d in departments]
    didx = _pick_index_numbered(console, f"2) Choisis un departement - {region}", department_labels, allow_back=True)
    if didx is None:
        return main(argv)
    department = departments[didx]

    city_presets = presets_by_department[department]
    table = Table(title=f"3) Villes - {region} / {department}", show_lines=False, header_style="bold magenta")
    table.add_column("#", justify="right", style="bold")
    table.add_column("Statut", justify="left")
    table.add_column("Ville", style="bold")
    table.add_column("Département", style="dim")
    table.add_column("Lignes (si dispo)", justify="right", style="dim")

    for i, pr in enumerate(city_presets, start=1):
        tag, style = _status_for_preset(out_root, pr)
        xlsx = _city_dir(out_root, pr) / f"{_safe_filename(pr.city)}.xlsx"
        nrows: Optional[int] = None
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(xlsx, read_only=True, data_only=True) if xlsx.exists() else None
            if wb is not None:
                try:
                    ws = wb.active
                    n = 0
                    for _ in ws.iter_rows(min_row=2, values_only=True):
                        n += 1
                    nrows = n
                finally:
                    wb.close()
        except Exception:
            nrows = None

        table.add_row(
            str(i),
            f"[{style}]{tag}[/{style}]",
            pr.city,
            pr.department or "-",
            str(nrows) if nrows is not None else "-",
        )

    console.print(table)
    while True:
        raw = console.input("Choix ville (numero, b=retour, q=quitter): ").strip().lower()
        if raw == "q":
            return 0
        if raw == "b":
            return main(argv)
        try:
            n = int(raw)
        except Exception:
            console.print("Entrée invalide.", style="yellow")
            continue
        if 1 <= n <= len(city_presets):
            preset = city_presets[n - 1]
            break
        console.print("Hors plage.", style="yellow")

    levels = ["centre", "agglo", "large", "tres_large"]
    lvl_table = Table(title=f"4) Level - {preset.city}", header_style="bold magenta")
    lvl_table.add_column("#", justify="right", style="bold")
    lvl_table.add_column("Level", style="bold")
    lvl_table.add_column("radius (m)", justify="right")
    lvl_table.add_column("step (m)", justify="right")
    lvl_table.add_column("rings", justify="right")
    for i, lv in enumerate(levels, start=1):
        v = preset.levels.get(lv) or {}
        if lv == "tres_large":
            base = preset.levels.get("large") or preset.levels.get("agglo") or preset.levels.get("centre") or {}
            v = {
                "radius": int(base.get("radius", 4000)) * 2,
                "grid_step_m": int(base.get("grid_step_m", 4000)) * 2,
                "grid_rings": max(4, int(base.get("grid_rings", 3)) + 1),
            }
        lvl_table.add_row(
            str(i),
            lv + ("  [dim](defaut)[/]" if lv == args.level else ""),
            str(v.get("radius", "-")),
            str(v.get("grid_step_m", "-")),
            str(v.get("grid_rings", "-")),
        )
    console.print(lvl_table)
    while True:
        raw = console.input("Choix level (1/2/3/4, Entree=defaut, q=quitter): ").strip().lower()
        if raw == "q":
            return 0
        if raw == "":
            level = args.level
            break
        try:
            n = int(raw)
        except Exception:
            console.print("Entrée invalide.", style="yellow")
            continue
        if 1 <= n <= len(levels):
            level = levels[n - 1]
            break
        console.print("Hors plage.", style="yellow")
    if level not in preset.levels and level != "tres_large":
        raise SystemExit(f"Level '{level}' non défini pour {preset.city}.")

    city_dir = _city_dir(out_root, preset)
    city_dir.mkdir(parents=True, exist_ok=True)
    city_parent_dir = city_dir.parent

    cache_path = Path(args.cache).expanduser().resolve() if args.cache else (city_dir / "places_details_cache.jsonl")

    cmd = _build_command(
        here=here,
        out_root=city_parent_dir,
        groups_file=groups_file,
        preset=preset,
        level=level,
        mode=args.mode,
        language=str(args.language),
        region_code=str(args.region_code) if args.region_code else None,
        timeout_s=int(args.timeout),
        limit_per_sector=int(args.limit_per_sector),
        details_max=int(args.details_max),
        cache_path=cache_path,
        category_translate=str(args.category_translate),
    )

    console.print(Panel.fit("5) Récap & lancement", style="bold cyan"))
    console.print(f"[bold]Ville[/]: {preset.city}  [dim]({preset.department})[/]")
    console.print(f"[bold]Level[/]: {level}   [bold]Mode[/]: {args.mode}")
    console.print(f"[bold]Cat translate[/]: {args.category_translate}")
    console.print(f"[bold]Sortie[/]: {city_dir}")
    console.print()
    console.print("[dim]Commande:[/]")
    console.print(" ".join([f"\"{c}\"" if " " in c else c for c in cmd]), style="dim")
    console.print()

    if args.dry_run:
        return 0

    go = console.input("Lancer l'export ? (o/n) [o] ").strip().lower() or "o"
    if go not in {"o", "oui", "y", "yes"}:
        console.print("Annulé.")
        return 0

    proc = subprocess.run(cmd)
    return int(proc.returncode or 0)


if __name__ == "__main__":
    raise SystemExit(main())

