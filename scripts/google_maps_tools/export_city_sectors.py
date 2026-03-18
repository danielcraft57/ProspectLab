import argparse
import io
import json
import shutil
import sys
import time
from contextlib import redirect_stdout
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

EXPORT_COLUMNS = [
    "name",
    "website",
    "category",
    "phone_number",
    "country",
    "address_1",
    "address_2",
    "longitude",
    "latitude",
    "rating",
    "reviews_count",
    "category_translate",
]


def _read_lines(path: str) -> List[str]:
    with open(path, "r", encoding="utf-8") as f:
        lines = [x.strip() for x in f.readlines()]
    return [x for x in lines if x and not x.startswith("#")]


def _safe_filename(name: str) -> str:
    keep: List[str] = []
    for ch in name.strip():
        if ch.isalnum():
            keep.append(ch)
        elif ch in {" ", "-", "_"}:
            keep.append("_")
    out = "".join(keep)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "secteur"


def _load_rows_from_xlsx(path: Path) -> List[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl n'est pas installe. Installe-le avec: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []

        header_map = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}
        out: List[dict] = []
        for row in rows_iter:
            obj = {}
            for col in EXPORT_COLUMNS:
                idx = header_map.get(col)
                obj[col] = row[idx] if idx is not None and idx < len(row) else None
            out.append(obj)
        return out
    finally:
        wb.close()


def _write_xlsx(path: Path, rows: List[dict]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl n'est pas installe. Installe-le avec: pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "export"
    ws.append(EXPORT_COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in EXPORT_COLUMNS])
    wb.save(path)


def _cleanup_group_dirs(base_dir: Path, group_name: str) -> None:
    target_slug = _safe_filename(group_name).lower()
    removed: List[str] = []
    for child in base_dir.iterdir():
        if not child.is_dir():
            continue
        if child.name.lower() == target_slug:
            for attempt in range(3):
                try:
                    shutil.rmtree(child)
                except Exception as e:
                    if attempt == 2:
                        sys.stdout.write(f"Impossible de supprimer le dossier '{child}': {e}\n")
                    else:
                        time.sleep(1.0)
                        continue
                break
            if not child.exists():
                removed.append(child.name)
    if removed:
        sys.stdout.write(f"Dossiers supprimes pour le groupe '{group_name}': {', '.join(removed)}\n\n")


def _read_groups_file(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    groups = data.get("groups") or []
    if not isinstance(groups, list) or not groups:
        raise SystemExit("Fichier groups invalide (groups manquant).")
    return groups


def _try_parse_last_json(stdout_text: str) -> Optional[Dict[str, Any]]:
    lines = [x.rstrip() for x in stdout_text.splitlines() if x.strip()]
    for i in range(len(lines) - 1, -1, -1):
        line = lines[i].strip()
        if line.endswith("}") and line.startswith("{"):
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    return obj
            except Exception:
                continue
    text = stdout_text.strip()
    start = text.rfind("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        chunk = text[start : end + 1]
        try:
            obj = json.loads(chunk)
            if isinstance(obj, dict):
                return obj
        except Exception:
            return None
    return None


def _run_batch_script(batch_script: str, cmd_args: List[str]) -> Tuple[int, str, Optional[Dict[str, Any]]]:
    buf = io.StringIO()
    rc = 0
    try:
        import runpy

        old_argv = sys.argv[:]
        sys.argv = cmd_args
        with redirect_stdout(buf):
            runpy.run_path(batch_script, run_name="__main__")
        sys.argv = old_argv
    except SystemExit as e:
        code = getattr(e, "code", 1)
        if isinstance(code, int):
            rc = code
        else:
            buf.write(str(code))
            buf.write("\n")
            rc = 1
    except Exception as e:
        buf.write(str(e))
        buf.write("\n")
        rc = 1

    out = buf.getvalue()
    return rc, out, _try_parse_last_json(out)


def main(argv: Optional[List[str]] = None) -> int:
    here = Path(__file__).resolve().parent
    default_sectors = str(here / "secteurs_economiques.txt")
    batch_script = str(here / "places_batch_export.py")

    p = argparse.ArgumentParser(description="Exporte 1 Excel par secteur dans un dossier par ville.")
    p.add_argument("--city", required=True, help="Ville (ex: Metz)")
    p.add_argument("--out-dir", required=True, help="Dossier racine de sortie (ex: Data)")
    p.add_argument("--sectors-file", default=default_sectors, help="Fichier secteurs (1 par ligne)")
    p.add_argument(
        "--groups-file",
        help="Optionnel: JSON de groupes (base: groupes_secteurs_grands.json). Si present, on exporte par groupe.",
    )
    p.add_argument("--mode", default="text", choices=["text", "nearby-grid"], help="Mode de recherche")
    p.add_argument("--lat", type=float, help="Latitude (nearby-grid)")
    p.add_argument("--lng", type=float, help="Longitude (nearby-grid)")
    p.add_argument("--radius", type=int, help="Rayon en metres (nearby-grid)")
    p.add_argument("--grid-step-m", type=int, default=1500, help="Pas de la grille en metres (nearby-grid)")
    p.add_argument("--grid-rings", type=int, default=3, help="Nombre d'anneaux (nearby-grid)")
    p.add_argument("--language", default="fr", help="Langue des champs")
    p.add_argument("--region", help="Code region (ex: fr)")
    p.add_argument("--timeout", type=int, default=25, help="Timeout reseau (secondes)")
    p.add_argument("--limit-per-sector", type=int, default=200, help="Nb max de lignes par secteur (website only)")
    p.add_argument("--details-max", type=int, default=200, help="Budget details par secteur (controle du cout)")
    p.add_argument("--cache", help="Cache JSONL. Defaut: out-dir/city/places_details_cache.jsonl")
    p.add_argument(
        "--category-translate",
        default="builtin",
        choices=["none", "builtin"],
        help="Traduction de category",
    )

    args = p.parse_args(argv)

    city = args.city.strip()
    if not city:
        raise SystemExit("Ville vide.")

    out_root = Path(args.out_dir).expanduser().resolve()
    out_city_dir = out_root / city
    out_city_dir.mkdir(parents=True, exist_ok=True)

    sectors = _read_lines(args.sectors_file)
    if not sectors:
        raise SystemExit("Aucun secteur dans le fichier secteurs.")

    cache_path = args.cache or str(out_city_dir / "places_details_cache.jsonl")

    exported = 0
    exported_files: List[Path] = []
    group_export_files: List[Path] = []
    errors: List[str] = []
    t0 = time.time()

    groups_mode = bool(args.groups_file)
    if groups_mode:
        groups_path = Path(str(args.groups_file)).expanduser().resolve()
        groups = _read_groups_file(groups_path)
        total_units = sum(len((g.get("sectors") or [])) for g in groups if isinstance(g, dict))
        sys.stdout.write(f"Ville: {city}\n")
        sys.stdout.write(f"Groupes: {len(groups)}\n")
        sys.stdout.write(f"Secteurs (total): {total_units}\n")
    else:
        sys.stdout.write(f"Ville: {city}\n")
        sys.stdout.write(f"Secteurs: {len(sectors)}\n")
    sys.stdout.write(f"Mode: {args.mode}\n")
    sys.stdout.write(f"Budget details max / secteur: {int(args.details_max)}\n")
    sys.stdout.write(f"Limit export / secteur: {int(args.limit_per_sector)}\n")
    sys.stdout.write("\n")

    total_places_seen = 0
    total_exported = 0
    total_details = 0

    jobs: List[Tuple[Optional[str], str]] = []
    if groups_mode:
        for g in groups:
            if not isinstance(g, dict):
                continue
            group_name = str(g.get("group") or "").strip()
            sec_list = g.get("sectors") or []
            if not group_name or not isinstance(sec_list, list):
                continue
            for s in sec_list:
                sector_clean = str(s or "").strip()
                if sector_clean:
                    jobs.append((group_name, sector_clean))
    else:
        for s in sectors:
            sector_clean = str(s or "").strip()
            if sector_clean:
                jobs.append((None, sector_clean))

    current_group: Optional[str] = None
    group_sector_files: List[Path] = []
    current_group_dir: Optional[Path] = None

    for group_name, sector_clean in jobs:
        if group_name != current_group:
            if current_group and group_sector_files:
                try:
                    merged_rows: List[dict] = []
                    for fp in group_sector_files:
                        if fp.exists():
                            merged_rows.extend(_load_rows_from_xlsx(fp))
                    seen = set()
                    deduped: List[dict] = []
                    for r in merged_rows:
                        key = (
                            str(r.get("website") or "").strip().lower(),
                            str(r.get("name") or "").strip().lower(),
                            str(r.get("address_1") or "").strip().lower(),
                        )
                        if key in seen:
                            continue
                        seen.add(key)
                        deduped.append(r)
                    out_group = out_city_dir / f"{_safe_filename(current_group)}.xlsx"
                    _write_xlsx(out_group, deduped)
                    sys.stdout.write(
                        f"Groupe termine: {current_group} -> {out_group} "
                        f"(brut: {len(merged_rows)}, apres dedup: {len(deduped)})\n\n"
                    )
                    group_export_files.append(out_group)

                    _cleanup_group_dirs(out_city_dir, current_group)
                except Exception as e:
                    sys.stdout.write(f"Impossible d'ecrire le fichier groupe {current_group}: {e}\n\n")

            current_group = group_name
            group_sector_files = []
            current_group_dir = None
            if current_group:
                sys.stdout.write(f"=== Groupe: {current_group} ===\n")

        if group_name:
            group_dir = out_city_dir / _safe_filename(group_name)
            group_dir.mkdir(parents=True, exist_ok=True)
            current_group_dir = group_dir
            out_xlsx = group_dir / f"{_safe_filename(sector_clean)}.xlsx"
        else:
            out_xlsx = out_city_dir / f"{_safe_filename(sector_clean)}.xlsx"

        cmd_args: List[str] = [
            batch_script,
            "--query",
            sector_clean,
            "--city",
            city,
            "--out",
            str(out_xlsx),
            "--limit",
            str(int(args.limit_per_sector)),
            "--details-max",
            str(int(args.details_max)),
            "--cache",
            cache_path,
            "--category-translate",
            str(args.category_translate),
            "--mode",
            str(args.mode),
            "--language",
            str(args.language),
            "--timeout",
            str(int(args.timeout)),
        ]
        if args.region:
            cmd_args.extend(["--region", str(args.region)])

        if args.mode == "nearby-grid":
            if args.lat is None or args.lng is None or args.radius is None:
                raise SystemExit("Mode nearby-grid: donne --lat, --lng et --radius.")
            cmd_args.extend(
                [
                    "--lat",
                    str(float(args.lat)),
                    "--lng",
                    str(float(args.lng)),
                    "--radius",
                    str(int(args.radius)),
                    "--grid-step-m",
                    str(int(args.grid_step_m)),
                    "--grid-rings",
                    str(int(args.grid_rings)),
                ]
            )

        idx = exported + len(errors) + 1
        sys.stdout.write(f"[{idx}/{len(jobs)}] Secteur: {sector_clean}\n")
        sector_t0 = time.time()
        rc, raw_out, summary = _run_batch_script(batch_script, cmd_args)
        elapsed = time.time() - sector_t0

        if rc == 0:
            exported += 1
            exported_files.append(out_xlsx)
            if group_name:
                group_sector_files.append(out_xlsx)
            if summary:
                places_seen = int(summary.get("places_seen") or 0)
                exported_n = int(summary.get("exported") or 0)
                details_used = int(summary.get("details_used") or 0)
                total_places_seen += places_seen
                total_exported += exported_n
                total_details += details_used
                sys.stdout.write(
                    f"  OK - places: {places_seen} | websites exportes: {exported_n} | details: {details_used} | {elapsed:.1f}s\n"
                )
            else:
                sys.stdout.write(f"  OK - {elapsed:.1f}s\n")
        else:
            msg = raw_out.strip() or "Erreur inconnue"
            errors.append(f"{sector_clean}: {msg}")
            sys.stdout.write(f"  ERREUR - {elapsed:.1f}s\n")
            sys.stdout.write(f"  {msg.splitlines()[-1]}\n")
            if "OVER_QUERY_LIMIT" in msg or "RESOURCE_EXHAUSTED" in msg:
                sys.stdout.write("  Quota Google detecte. On stoppe ici pour eviter de boucler inutilement.\n")
                break

        sys.stdout.write("\n")

    if groups_mode and current_group and group_sector_files:
        try:
            merged_rows = []
            for fp in group_sector_files:
                if fp.exists():
                    merged_rows.extend(_load_rows_from_xlsx(fp))
            seen = set()
            deduped = []
            for r in merged_rows:
                key = (
                    str(r.get("website") or "").strip().lower(),
                    str(r.get("name") or "").strip().lower(),
                    str(r.get("address_1") or "").strip().lower(),
                )
                if key in seen:
                    continue
                seen.add(key)
                deduped.append(r)
            out_group = out_city_dir / f"{_safe_filename(current_group)}.xlsx"
            _write_xlsx(out_group, deduped)
            sys.stdout.write(
                f"Groupe termine: {current_group} -> {out_group} "
                f"(brut: {len(merged_rows)}, apres dedup: {len(deduped)})\n"
            )
            group_export_files.append(out_group)

            _cleanup_group_dirs(out_city_dir, current_group)
        except Exception as e:
            sys.stdout.write(f"Impossible d'ecrire le fichier groupe {current_group}: {e}\n\n")

    try:
        merged_rows: List[dict] = []
        # En mode groupes, les fichiers secteurs sont supprimes. On reconstruit donc le global
        # a partir des fichiers de groupes.
        sources = group_export_files if groups_mode else exported_files
        for fpath in sources:
            if fpath.exists():
                merged_rows.extend(_load_rows_from_xlsx(fpath))

        seen = set()
        deduped: List[dict] = []
        for r in merged_rows:
            key = (
                str(r.get("website") or "").strip().lower(),
                str(r.get("name") or "").strip().lower(),
                str(r.get("address_1") or "").strip().lower(),
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)

        out_all = out_city_dir / f"{_safe_filename(city)}.xlsx"
        _write_xlsx(out_all, deduped)
        sys.stdout.write(
            f"Fichier global cree: {out_all} (brut: {len(merged_rows)}, apres dedup: {len(deduped)})\n"
        )
    except Exception as e:
        sys.stdout.write(f"Impossible de creer le fichier global: {e}\n")

    total_elapsed = time.time() - t0
    sys.stdout.write("\nResume\n")
    sys.stdout.write(f"- Secteurs OK: {exported}\n")
    sys.stdout.write(f"- Secteurs en erreur: {len(errors)}\n")
    sys.stdout.write(f"- Places vues (total): {total_places_seen}\n")
    sys.stdout.write(f"- Websites exportes (total): {total_exported}\n")
    sys.stdout.write(f"- Details consommes (total): {total_details}\n")
    sys.stdout.write(f"- Temps total: {total_elapsed/60.0:.1f} min\n")
    sys.stdout.write(f"- Dossier: {out_city_dir}\n")

    if errors:
        sys.stdout.write("\nErreurs\n")
        for e in errors[:20]:
            sys.stdout.write(f"- {e.splitlines()[0]}\n")
        if len(errors) > 20:
            sys.stdout.write(f"- ... ({len(errors) - 20} autres)\n")

    sys.stdout.write(f"\nExports termines pour {exported} secteurs dans: {out_city_dir}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


