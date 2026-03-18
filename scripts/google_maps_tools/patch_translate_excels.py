import argparse
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple


def _load_translator():
    # Import local (même dossier) sans dépendre de PYTHONPATH.
    here = Path(__file__).resolve().parent
    sys.path.insert(0, str(here))
    try:
        import places_search as ps  # type: ignore

        return ps._translate_category  # noqa: SLF001
    finally:
        try:
            sys.path.remove(str(here))
        except Exception:
            pass


def _safe_filename(name: str) -> str:
    keep: List[str] = []
    for ch in name.strip():
        if ch.isalnum():
            keep.append(ch)
        elif ch in {" ", "-", "_"}:
            keep.append("_")
    out = "".join(keep)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "fichier"


def _iter_xlsx(root: Path) -> List[Path]:
    return sorted([p for p in root.rglob("*.xlsx") if p.is_file()])


def _openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit("openpyxl n'est pas installe. Installe-le avec: pip install openpyxl") from e
    return load_workbook


def _header_map(header_row) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip()
        if key:
            out[key] = idx
    return out


def _fmt_s(seconds: float) -> str:
    s = max(0, int(seconds))
    h = s // 3600
    m = (s % 3600) // 60
    sec = s % 60
    if h:
        return f"{h}h{m:02d}m"
    if m:
        return f"{m}m{sec:02d}s"
    return f"{sec}s"


def _fmt_pct(x: float) -> str:
    try:
        return f"{x:.1f}%"
    except Exception:
        return "?"


def _patch_one(
    xlsx: Path,
    *,
    translate_fn,
    in_place: bool,
    backup: bool,
    set_category: bool,
    verbose: bool,
) -> Tuple[int, int]:
    load_workbook = _openpyxl()
    wb = load_workbook(xlsx)
    try:
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        hm = _header_map(header)

        if "category" not in hm:
            return 0, 0

        cat_idx = hm["category"] + 1
        # Si la colonne n'existe pas, on l'ajoute en fin.
        if "category_translate" in hm:
            cat_tr_idx = hm["category_translate"] + 1
        else:
            cat_tr_idx = ws.max_column + 1
            ws.cell(row=1, column=cat_tr_idx, value="category_translate")

        changed = 0
        total = 0
        for row in range(2, ws.max_row + 1):
            total += 1
            raw_cat = ws.cell(row=row, column=cat_idx).value
            if raw_cat is None:
                continue
            raw_cat_s = str(raw_cat).strip()
            if not raw_cat_s:
                continue
            new_tr = translate_fn(raw_cat_s, mode="builtin")
            if new_tr is None:
                continue
            old_tr = ws.cell(row=row, column=cat_tr_idx).value
            old_tr_s = (str(old_tr).strip() if old_tr is not None else "")
            if old_tr_s != str(new_tr):
                ws.cell(row=row, column=cat_tr_idx, value=str(new_tr))
                changed += 1
            if set_category and raw_cat_s != str(new_tr):
                ws.cell(row=row, column=cat_idx, value=str(new_tr))

        if changed == 0 and not set_category:
            return total, 0

        if in_place:
            if backup:
                bak = xlsx.with_suffix(xlsx.suffix + ".bak")
                if not bak.exists():
                    try:
                        xlsx.replace(bak)
                    except Exception:
                        # fallback: si replace échoue, on n'empêche pas le patch.
                        pass
                    else:
                        # On sauvegarde ensuite sur le nom original
                        out_path = xlsx
                        wb.save(out_path)
                        return total, changed
            wb.save(xlsx)
        else:
            out_path = xlsx.with_name(f"{xlsx.stem}.patched.xlsx")
            wb.save(out_path)
            if verbose:
                print(f"  -> {out_path}")

        return total, changed
    finally:
        wb.close()


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Patch les fichiers Excel exportés pour recalculer category_translate (et optionnellement category)."
    )
    p.add_argument("--root", default="Data", help="Dossier racine (ex: Data)")
    p.add_argument("--city", help="Optionnel: limiter à une ville (Data/<Ville>/...)")
    p.add_argument("--in-place", action="store_true", help="Modifie les fichiers en place (sinon: *.patched.xlsx)")
    p.add_argument("--no-backup", action="store_true", help="En in-place, ne pas créer de .bak")
    p.add_argument("--set-category", action="store_true", help="Remplace aussi la colonne category par la traduction")
    p.add_argument("--verbose", action="store_true", help="Plus de détails")
    p.add_argument(
        "--progress-every",
        type=int,
        default=1,
        help="Affiche une ligne d'avancement toutes les N fichiers (defaut: 1)",
    )
    args = p.parse_args(argv)

    root = Path(str(args.root)).expanduser().resolve()
    if args.city:
        root = root / str(args.city).strip()
    if not root.exists():
        raise SystemExit(f"Dossier introuvable: {root}")

    translate_fn = _load_translator()
    files = _iter_xlsx(root)
    if not files:
        print("Aucun .xlsx trouvé.")
        return 0

    in_place = bool(args.in_place)
    backup = not bool(args.no_backup)
    total_rows = 0
    total_changed = 0
    total_files = 0

    nfiles = len(files)
    progress_every = max(1, int(args.progress_every))
    t0 = time.time()
    last_print = 0

    print(f"Fichiers: {nfiles} | in_place={in_place} | backup={backup} | set_category={bool(args.set_category)}")
    for i, fp in enumerate(files, start=1):
        total_files += 1
        try:
            rows, changed = _patch_one(
                fp,
                translate_fn=translate_fn,
                in_place=in_place,
                backup=backup,
                set_category=bool(args.set_category),
                verbose=bool(args.verbose),
            )
            total_rows += rows
            total_changed += changed
            if args.verbose:
                print(f"- {fp}: {changed} modifs")
        except Exception as e:
            print(f"- ERREUR {fp}: {e}")

        # Avancement (1 ligne)
        if (i - last_print) >= progress_every or i == nfiles:
            elapsed = time.time() - t0
            rate = (i / elapsed) if elapsed > 0 else 0.0
            remaining = (nfiles - i) / rate if rate > 0 else 0.0
            pct = (i * 100.0 / nfiles) if nfiles else 100.0
            print(
                f"[{i}/{nfiles}] {_fmt_pct(pct)} | modifs: {total_changed} | lignes: {total_rows} | "
                f"elapsed: {_fmt_s(elapsed)} | eta: {_fmt_s(remaining)}"
            )
            last_print = i

    total_elapsed = time.time() - t0
    print(
        f"Terminé. Fichiers traités: {total_files} | Lignes: {total_rows} | Modifs: {total_changed} | "
        f"Temps: {_fmt_s(total_elapsed)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

