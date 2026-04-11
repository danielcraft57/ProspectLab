"""
Export Excel multi-feuilles pour la vue « Analyse concurrence & marché ».
"""

from __future__ import annotations

from io import BytesIO
from typing import Any


def _sheet_dict_counts(ws, title: str, data: dict[str, Any], key_header: str, count_header: str = 'Effectifs') -> None:
    ws.append([title])
    ws.append([key_header, count_header])
    items = sorted((data or {}).items(), key=lambda kv: (-(kv[1] or 0), (kv[0] or '').lower()))
    for k, v in items:
        ws.append([k, int(v or 0)])


def _sheet_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append(r)


def build_market_concurrence_xlsx(stats: dict[str, Any]) -> BytesIO:
    """
    Construit un classeur .xlsx à partir du JSON renvoyé par get_statistics().

    Args:
        stats: dictionnaire statistiques (incl. par_pays, geo_resume, evolution_trimestrielle, etc.)

    Returns:
        BytesIO positionné au début, prêt pour send_file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl est requis pour l’export Excel.") from e

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    bold = Font(bold=True)

    # --- Synthèse ---
    ws0 = wb.create_sheet('Synthèse', 0)
    ws0.append(['Indicateur', 'Valeur'])
    for c in ws0[1]:
        c.font = bold
    lines = [
        ('Total entreprises', stats.get('total_entreprises', 0)),
        ('Entreprises avec email', stats.get('entreprises_avec_email', 0)),
        ('Analyses totales', stats.get('total_analyses', 0)),
        ('Favoris', stats.get('favoris', 0)),
        ('Emails envoyés (période filtre API)', stats.get('emails_envoyes', 0)),
        ('Taux ouverture %', stats.get('open_rate', 0)),
        ('Taux clic %', stats.get('click_rate', 0)),
        ('Campagnes (période)', stats.get('total_campagnes', 0)),
    ]
    geo = stats.get('geo_resume') or {}
    if geo:
        lines.extend([
            ('GPS renseigné (lat+lon)', geo.get('avec_coords', 0)),
            ('Sans GPS', geo.get('sans_coords', 0)),
            ('GPS approx. France métropole', geo.get('france_metropole_approx', 0)),
        ])
    for label, val in lines:
        ws0.append([label, val])

    # --- Opportunités ---
    ws1 = wb.create_sheet('Opportunités')
    _sheet_dict_counts(ws1, 'Répartition par opportunité', stats.get('par_opportunite') or {}, 'Opportunité')

    # --- Statuts ---
    ws2 = wb.create_sheet('Statuts')
    _sheet_dict_counts(ws2, 'Répartition par statut', stats.get('par_statut') or {}, 'Statut')

    # --- Secteurs ---
    ws3 = wb.create_sheet('Secteurs')
    _sheet_dict_counts(ws3, 'Répartition par secteur', stats.get('par_secteur') or {}, 'Secteur')

    # --- Pays ---
    ws4 = wb.create_sheet('Pays')
    _sheet_dict_counts(ws4, 'Répartition par pays', stats.get('par_pays') or {}, 'Pays')

    # --- Étapes prospection ---
    ws5 = wb.create_sheet('Étapes prospection')
    _sheet_dict_counts(ws5, 'Étapes', stats.get('par_etape_prospection') or {}, 'Étape')

    # --- Trimestres ---
    ws6 = wb.create_sheet('Évolution trimestrielle')
    ws6.append(['Période (année-Ttrimestre)', 'Nouvelles fiches (date_analyse)', 'Gagnés (sur ce trimestre)'])
    for c in ws6[1]:
        c.font = bold
    evo = list(stats.get('evolution_trimestrielle') or [])
    evo_sorted = sorted(evo, key=lambda x: (x.get('periode') or ''))
    for row in evo_sorted:
        ws6.append([
            row.get('periode'),
            int(row.get('nouvelles') or 0),
            int(row.get('gagnes') or 0),
        ])

    # --- Priorités ---
    ws7 = wb.create_sheet('Priorités')
    ws7.append(['id', 'nom', 'secteur', 'opportunite', 'statut', 'website'])
    for c in ws7[1]:
        c.font = bold
    for p in stats.get('priority_prospects') or []:
        ws7.append([
            p.get('id'),
            p.get('nom'),
            p.get('secteur'),
            p.get('opportunite'),
            p.get('statut'),
            p.get('website'),
        ])

    # --- Tags ---
    ws8 = wb.create_sheet('Tags')
    ws8.append(['Tag', 'Effectifs'])
    for c in ws8[1]:
        c.font = bold
    for t in stats.get('top_tags') or []:
        ws8.append([t.get('tag'), int(t.get('count') or 0)])

    # --- Campagnes récentes ---
    ws9 = wb.create_sheet('Campagnes récentes')
    ws9.append(['id', 'nom', 'statut', 'destinataires', 'envoyés', 'réussis', 'date_creation'])
    for c in ws9[1]:
        c.font = bold
    for c in stats.get('recent_campagnes') or []:
        ws9.append([
            c.get('id'),
            c.get('nom'),
            c.get('statut'),
            c.get('total_destinataires'),
            c.get('total_envoyes'),
            c.get('total_reussis'),
            str(c.get('date_creation') or ''),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
